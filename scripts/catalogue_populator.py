#!/usr/bin/env python3
"""Scan Teaching Resources OneDrive folder and build the Excel catalogue."""

import sys
import json
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook

# ── Paths ────────────────────────────────────────────────────────────────────
TEACHING_ROOT = Path(
    r"C:\Users\dan_m\OneDrive\Education\PGDip Science AI for Business"
    r"\S3 02 Customer Engagement and Artificial Intelligence\CA3 Final Project"
    r"\Teaching Resources"
)
OUTPUT_XLSX = Path(
    r"C:\Users\dan_m\OneDrive\Education\PGDip Science AI for Business"
    r"\S3 02 Customer Engagement and Artificial Intelligence\CA3 Final Project"
    r"\GitRepo\scripts\catalogue_output.xlsx"
)
SKIP_FOLDER = "Elaine Sony Vaio"

# ── Folder-to-subject mapping ────────────────────────────────────────────────
# Keys are lowercase folder names (or partial patterns ending with *)
SUBJECT_MAP = {
    # Literacy
    "phonics": "Literacy",
    "morphology": "Literacy",
    "spelling": "Literacy",
    "writing": "Literacy",
    "comprehension": "Literacy",
    "grammar-punct": "Literacy",
    "decodeables": "Literacy",
    "word recogn": "Literacy",
    "poem": "Literacy",
    "nursery rhyme": "Literacy",
    "syllables": "Literacy",
    "alphabet order": "Literacy",
    "sentence pyramids fluency": "Literacy",
    "language": "Literacy",
    "novel study": "Literacy",
    "readingposters": "Literacy",
    "roald dahl": "Literacy",
    "littleredridinghood": "Literacy",
    "retelling story": "Literacy",
    "sor": "Literacy",
    # Maths
    "maths": "Maths",
    # Gaeilge
    "gaeilge": "Gaeilge",
    # SESE
    "sese": "SESE",
    # SPHE
    "sel": "SPHE",
    "bd": "SPHE",
    # Arts
    "art": "Arts",
    "music": "Arts",
    # PE
    "pe_movement": "PE",
    "gross motor": "PE",
    "fine motor": "PE",
    # Religion
    "religion": "Religion",
}

# Special prefixes that match from the start (e.g. "SOR Pack" matches "sor")
SUBJECT_PREFIX_MAP = {}
for key, val in SUBJECT_MAP.items():
    SUBJECT_PREFIX_MAP[key] = val

# Extra pattern-based mappings (checked with startswith)
SUBJECT_PATTERNS = [
    ("sor ", "Literacy"),
    ("pr-4037", "SESE"),
    ("ngk-", "SESE"),
    ("anti-cyber", "SPHE"),
]

# Filename-specific overrides for root-level PDFs
FILENAME_SUBJECT_MAP = {
    "math toolkit bw.pdf": "Maths",
    "exercise-bingo.pdf": "SPHE",
    "fitness-circuit-printables.pdf": "SPHE",
}

# ── Season detection ─────────────────────────────────────────────────────────
SEASON_MAP = {
    "xmas": "Christmas",
    "halloween": "Halloween",
    "easter": "Easter",
    "stpat": "St Patricks Day",
    "valentine": "Valentines",
    "spring": "Spring",
    "summer": "Summer",
    "back to school": "Back-to-School",
    "end year": "End-of-Year",
    "pancakes": "Pancake Tuesday",
    "leap day": "Leap Day",
    "fall resources literacy numeracy": "Autumn",
}

# ── Non-PDF extensions to skip ───────────────────────────────────────────────
SKIP_EXTENSIONS = {".pptx", ".docx", ".jpg", ".jpeg", ".png", ".gif",
                   ".zip", ".mp4", ".mov", ".avi", ".ppt", ".doc",
                   ".xlsx", ".xls", ".pptm", ".docm", ".pub", ".pages",
                   ".key", ".numbers", ".bmp", ".tiff", ".webp", ".svg",
                   ".mp3", ".wav", ".wma", ".flv", ".mkv", ".webm"}

# ── Column names ─────────────────────────────────────────────────────────────
COLUMNS = [
    "id",
    "filename",
    "onedrive_path",
    "subject",
    "subdomain",
    "grade_band",
    "format",
    "activity_type",
    "season",
    "pedagogy",
    "curriculum_area",
    "strand",
    "outcome_code",
    "programme",
    "programme_ref",
    "source",
    "license",
    "confidence",
    "ai_generated",
    "image_heavy",
    "extracted_text_chars",
    "extracted_text_sample",
]


def detect_subject_and_season(relative_parts):
    """Determine subject, subdomain, and season from folder path parts.

    relative_parts: list of folder names (no filename) relative to
    Teaching Resources root, e.g. ['phonics', 'digraphs'].
    """
    parts_lower = [p.lower() for p in relative_parts]

    # Detect season
    season = "Generic"
    for part in parts_lower:
        if part in SEASON_MAP:
            season = SEASON_MAP[part]
            break

    # Detect subject — prefer the deepest match
    subject = "Other"
    for part in parts_lower:
        if part in SUBJECT_MAP:
            subject = SUBJECT_MAP[part]
        else:
            for pattern, subj in SUBJECT_PATTERNS:
                if part.startswith(pattern):
                    subject = subj
                    break

    # Build subdomain from all folder parts (joined with +)
    subdomain_parts = [p.strip().title() for p in relative_parts]
    subdomain = "+".join(subdomain_parts) if subdomain_parts else ""

    return subject, subdomain, season


def detect_root_subject(filename_lower):
    """For root-level PDFs, try to detect subject from filename."""
    if filename_lower in FILENAME_SUBJECT_MAP:
        return FILENAME_SUBJECT_MAP[filename_lower]
    for pattern, subj in SUBJECT_PATTERNS:
        if filename_lower.startswith(pattern):
            return subj
    return "Other"


def extract_pdf_text(pdf_path):
    """Extract text from a PDF using pdfplumber. Returns (text, char_count)."""
    text_parts = []
    total_chars = 0
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
                    total_chars += len(page_text)
    except Exception as exc:
        print(f"  [WARN] Failed to read PDF: {pdf_path.name} — {exc}",
              file=sys.stderr)
        return "", 0

    full_text = "\n".join(text_parts)
    return full_text, total_chars


def load_existing_paths(output_path):
    """Load set of already-catalogued OneDrive paths from existing Excel."""
    if not output_path.exists():
        return set()
    try:
        wb = load_workbook(output_path, read_only=True)
        ws = wb.active
        existing = set()
        # Find the onedrive_path column index
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            col_idx = header.index("onedrive_path")
        except ValueError:
            wb.close()
            return set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                existing.add(str(val))
        wb.close()
        return existing
    except Exception as exc:
        print(f"  [WARN] Could not read existing catalogue: {exc}",
              file=sys.stderr)
        return set()


def write_catalogue(rows, output_path):
    """Write catalogue rows to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogue"

    # Header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # Auto-width rough pass
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    wb.save(output_path)
    print(f"\nCatalogue written to: {output_path}")


def main():
    print(f"Scanning: {TEACHING_ROOT}")
    print(f"Output:   {OUTPUT_XLSX}\n")

    # Load existing catalogue for idempotency
    existing_paths = load_existing_paths(OUTPUT_XLSX)
    print(f"Existing entries in catalogue: {len(existing_paths)}\n")

    # Collect all PDF files
    all_pdfs = []
    for entry in TEACHING_ROOT.rglob("*"):
        # Skip the excluded folder
        if SKIP_FOLDER in entry.parts:
            continue
        if entry.is_file() and entry.suffix.lower() == ".pdf":
            all_pdfs.append(entry)

    total_found = len(all_pdfs)
    print(f"Total PDF files found: {total_found}\n")

    rows = []
    skipped = 0
    processed = 0
    failed = 0
    total_chars_all = 0
    image_heavy_count = 0

    # Load previous max ID
    next_id = 1
    if existing_paths:
        # Try to find highest existing ID
        try:
            wb = load_workbook(OUTPUT_XLSX, read_only=True)
            ws = wb.active
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            id_col = header.index("id") if "id" in header else -1
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if id_col >= 0 and id_col < len(row):
                    val = row[id_col]
                    if val and isinstance(val, str) and val.startswith("RES-"):
                        try:
                            num = int(val[4:])
                            max_id = max(max_id, num)
                        except ValueError:
                            pass
            next_id = max_id + 1
            wb.close()
        except Exception:
            next_id = 1

    for idx, pdf_path in enumerate(all_pdfs):
        # Build the OneDrive path (relative to Teaching Resources root)
        try:
            onedrive_path = str(pdf_path.relative_to(TEACHING_ROOT))
        except ValueError:
            onedrive_path = pdf_path.name

        # Idempotency check
        if onedrive_path in existing_paths:
            skipped += 1
            continue

        # Extract folder parts (exclude filename)
        relative_parts = list(pdf_path.relative_to(TEACHING_ROOT).parts[:-1])

        # Determine subject, subdomain, season
        if relative_parts:
            subject, subdomain, season = detect_subject_and_season(relative_parts)
        else:
            subject = detect_root_subject(pdf_path.name.lower())
            subdomain = ""
            season = "Generic"
            # Also check filename for season hints
            fname_lower = pdf_path.name.lower()
            if "christmas" in fname_lower or "xmas" in fname_lower:
                season = "Christmas"
            elif "halloween" in fname_lower:
                season = "Halloween"
            elif "easter" in fname_lower:
                season = "Easter"
            elif "valentine" in fname_lower:
                season = "Valentines"
            elif "spring" in fname_lower:
                season = "Spring"
            elif "summer" in fname_lower:
                season = "Summer"

        # Extract text
        extracted_text, char_count = extract_pdf_text(pdf_path)
        if char_count == 0 and extracted_text == "":
            failed += 1
            continue  # skip entirely unreadable files

        image_heavy = char_count < 500
        text_sample = extracted_text[:5000] if extracted_text else ""

        row_id = f"RES-{next_id:04d}"
        next_id += 1

        rows.append({
            "id": row_id,
            "filename": pdf_path.name,
            "onedrive_path": onedrive_path,
            "subject": subject,
            "subdomain": subdomain,
            "grade_band": "",
            "format": "",
            "activity_type": "",
            "season": season,
            "pedagogy": "",
            "curriculum_area": "",
            "strand": "",
            "outcome_code": "",
            "programme": "",
            "programme_ref": "",
            "source": "",
            "license": "",
            "confidence": 0.5,
            "ai_generated": False,
            "image_heavy": image_heavy,
            "extracted_text_chars": char_count,
            "extracted_text_sample": text_sample,
        })

        total_chars_all += char_count
        if image_heavy:
            image_heavy_count += 1
        processed += 1

        if processed % 50 == 0:
            print(f"Processed {processed}/{total_found} files...")

    # ── Summary ──────────────────────────────────────────────────────────────
    avg_chars = total_chars_all / processed if processed > 0 else 0
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"  Total files found:             {total_found}")
    print(f"  Files skipped (in catalogue):   {skipped}")
    print(f"  Files processed:                {processed}")
    print(f"  Files failed:                   {failed}")
    print(f"  Average chars per file:         {avg_chars:,.0f}")
    print(f"  Files marked image_heavy:       {image_heavy_count}")
    print(f"  New catalogue entries:          {len(rows)}")

    if rows:
        write_catalogue(rows, OUTPUT_XLSX)
    else:
        print("\nNo new files to add — catalogue is up to date.")


if __name__ == "__main__":
    main()
